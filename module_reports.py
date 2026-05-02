"""
module_reports.py — 👤 ÜYE 4: Raporlama ve Analiz
═══════════════════════════════════════════════════════════════
Bu modül 4 katmanı da kapsar:
  🎨 GUI:    Tarih seçici, rapor türü dropdown, butonlar, önizleme
  💾 DB:     Aggregation sorguları (SUM, GROUP BY, ORDER BY)
  📊 Grafik: Heatmap (seaborn), violin plot (seaborn), bar chart (mpl)
  🧠 Mantık: Pandas DataFrame, KPI hesabı, .to_excel() / .to_csv()

Sunum İpucu:
  - "En Çok Satanlar" raporunu Excel'e aktar, dosyayı aç.
  - Heatmap'i göster: "Saatler × ürünler, hücre = ciro"
  - Hocaya: "Pandas + seaborn ile profesyonel analiz görselleri"
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import os

import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

import database
import config


# Seaborn stilini globalce ayarla
sns.set_style("whitegrid")


class ReportsModule(ttk.Frame):
    """Raporlama sekmesi."""

    REPORT_TYPES = [
        "En Çok Satan Ürünler",
        "Saatlik Ciro",
        "Ürün × Saat Heatmap",
        "Fiyat Dağılımı (Violin)",
        "Kategori Cirosu",
        "Genel Özet (KPI)",
    ]

    def __init__(self, parent):
        super().__init__(parent)
        self._build_ui()

    # ─────────────────────────────────────────────────────
    # UI
    # ─────────────────────────────────────────────────────
    def _build_ui(self):
        # Başlık
        ttk.Label(
            self, text="📊  Raporlama ve Analiz",
            font=config.FONT_TITLE
        ).pack(pady=(10, 5), padx=10, anchor="w")

        ttk.Label(
            self,
            text="Pandas + Matplotlib + Seaborn ile satış analizleri. "
                 "Excel/CSV'ye aktarın veya görselleştirin.",
            font=config.FONT_NORMAL
        ).pack(pady=(0, 10), padx=10, anchor="w")

        # ── Kontrol paneli ──
        control = ttk.LabelFrame(self, text=" Rapor Seçimi ")
        control.pack(fill="x", padx=10, pady=5)

        ttk.Label(control, text="Rapor Türü:").grid(row=0, column=0, padx=8, pady=8, sticky="w")
        self.report_var = tk.StringVar(value=self.REPORT_TYPES[0])
        report_dd = ttk.Combobox(
            control, textvariable=self.report_var,
            values=self.REPORT_TYPES, state="readonly", width=28
        )
        report_dd.grid(row=0, column=1, padx=4, pady=8)

        ttk.Label(control, text="Veri Kaynağı:").grid(row=0, column=2, padx=(20, 4), pady=8, sticky="w")
        self.source_var = tk.StringVar(value="Tümü")
        source_dd = ttk.Combobox(
            control, textvariable=self.source_var,
            values=["Tümü", "Sadece Gerçek", "Sadece Simülasyon"],
            state="readonly", width=18
        )
        source_dd.grid(row=0, column=3, padx=4, pady=8)

        # Butonlar
        ttk.Button(control, text="📊 Görselleştir",
                   command=self.show_visualization).grid(row=0, column=4, padx=4, pady=8)
        ttk.Button(control, text="📤 Excel'e Aktar",
                   command=self.export_excel).grid(row=0, column=5, padx=4, pady=8)
        ttk.Button(control, text="📥 CSV'ye Aktar",
                   command=self.export_csv).grid(row=0, column=6, padx=4, pady=8)

        # ── Ana alan: önizleme + grafik ──
        body = ttk.Frame(self)
        body.pack(fill="both", expand=True, padx=10, pady=5)
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=2)
        body.rowconfigure(0, weight=1)

        # Sol: önizleme tablosu (Treeview)
        preview_wrap = ttk.LabelFrame(body, text=" Veri Önizlemesi ")
        preview_wrap.grid(row=0, column=0, sticky="nsew", padx=(0, 10))

        self.preview_tree = ttk.Treeview(preview_wrap, show="headings", height=15)
        self.preview_tree.pack(fill="both", expand=True, padx=4, pady=4)

        # Sağ: grafik
        chart_wrap = ttk.LabelFrame(body, text=" Grafik ")
        chart_wrap.grid(row=0, column=1, sticky="nsew")

        self.fig = Figure(figsize=(7, 5), dpi=90)
        self.ax = self.fig.add_subplot(111)
        self.canvas = FigureCanvasTkAgg(self.fig, master=chart_wrap)
        self.canvas.get_tk_widget().pack(fill="both", expand=True, padx=4, pady=4)

        self._show_placeholder()

    def _show_placeholder(self):
        self.ax.clear()
        self.ax.text(0.5, 0.5,
                     "Rapor seçip 'Görselleştir' butonuna basın.",
                     ha="center", va="center",
                     transform=self.ax.transAxes,
                     fontsize=11, color="#94a3b8")
        self.ax.axis("off")
        self.canvas.draw_idle()

    # ─────────────────────────────────────────────────────
    # Veri Çekme
    # ─────────────────────────────────────────────────────
    def _source_filter(self) -> str:
        """SQL WHERE eki: 'Tümü' / 'Sadece Gerçek' / 'Sadece Simülasyon'."""
        s = self.source_var.get()
        if s == "Sadece Gerçek":
            return " AND o.is_simulated=0 "
        if s == "Sadece Simülasyon":
            return " AND o.is_simulated=1 "
        return ""

    def _get_dataframe(self) -> pd.DataFrame:
        """Seçili rapora göre veri çeker."""
        report = self.report_var.get()
        flt = self._source_filter()

        with database.get_connection() as conn:
            if report == "En Çok Satan Ürünler":
                return pd.read_sql_query(
                    f"SELECT p.name AS Ürün, p.category AS Kategori, "
                    f"       SUM(oi.quantity) AS Adet, "
                    f"       SUM(oi.quantity * oi.locked_price) AS Ciro "
                    f"FROM order_items oi "
                    f"JOIN orders o ON oi.order_id = o.id "
                    f"JOIN products p ON oi.product_id = p.id "
                    f"WHERE 1=1 {flt} "
                    f"GROUP BY p.id "
                    f"ORDER BY Ciro DESC", conn
                )

            elif report == "Saatlik Ciro":
                return pd.read_sql_query(
                    f"SELECT strftime('%H', o.created_at) AS Saat, "
                    f"       COUNT(o.id) AS Sipariş, "
                    f"       SUM(o.total_amount) AS Ciro "
                    f"FROM orders o "
                    f"WHERE 1=1 {flt} "
                    f"GROUP BY Saat ORDER BY Saat", conn
                )

            elif report == "Ürün × Saat Heatmap":
                return pd.read_sql_query(
                    f"SELECT p.name AS Ürün, "
                    f"       strftime('%H', o.created_at) AS Saat, "
                    f"       SUM(oi.quantity * oi.locked_price) AS Ciro "
                    f"FROM order_items oi "
                    f"JOIN orders o ON oi.order_id = o.id "
                    f"JOIN products p ON oi.product_id = p.id "
                    f"WHERE 1=1 {flt} "
                    f"GROUP BY p.name, Saat", conn
                )

            elif report == "Fiyat Dağılımı (Violin)":
                return pd.read_sql_query(
                    "SELECT p.name AS Ürün, ph.price AS Fiyat "
                    "FROM price_history ph "
                    "JOIN products p ON ph.product_id = p.id", conn
                )

            elif report == "Kategori Cirosu":
                return pd.read_sql_query(
                    f"SELECT p.category AS Kategori, "
                    f"       SUM(oi.quantity * oi.locked_price) AS Ciro, "
                    f"       SUM(oi.quantity) AS Adet "
                    f"FROM order_items oi "
                    f"JOIN orders o ON oi.order_id = o.id "
                    f"JOIN products p ON oi.product_id = p.id "
                    f"WHERE 1=1 {flt} "
                    f"GROUP BY p.category ORDER BY Ciro DESC", conn
                )

            elif report == "Genel Özet (KPI)":
                rows = []
                cur = conn.cursor()

                cur.execute(f"SELECT COUNT(*) FROM orders o WHERE 1=1 {flt}")
                rows.append(("Toplam Sipariş Sayısı", cur.fetchone()[0]))

                cur.execute(f"SELECT COALESCE(SUM(total_amount),0) FROM orders o WHERE 1=1 {flt}")
                rows.append(("Toplam Ciro (₺)", round(cur.fetchone()[0], 2)))

                cur.execute(f"SELECT COALESCE(AVG(total_amount),0) FROM orders o WHERE 1=1 {flt}")
                rows.append(("Ortalama Sepet (₺)", round(cur.fetchone()[0], 2)))

                cur.execute(
                    f"SELECT COALESCE(SUM(oi.quantity),0) FROM order_items oi "
                    f"JOIN orders o ON oi.order_id=o.id WHERE 1=1 {flt}"
                )
                rows.append(("Toplam Satılan Adet", cur.fetchone()[0]))

                cur.execute("SELECT COUNT(*) FROM products WHERE is_active=1")
                rows.append(("Aktif Ürün Sayısı", cur.fetchone()[0]))

                return pd.DataFrame(rows, columns=["Metrik", "Değer"])

        return pd.DataFrame()

    # ─────────────────────────────────────────────────────
    # Önizleme
    # ─────────────────────────────────────────────────────
    def _populate_preview(self, df: pd.DataFrame):
        # Treeview'i sıfırla
        for col in self.preview_tree["columns"]:
            self.preview_tree.heading(col, text="")
        self.preview_tree.delete(*self.preview_tree.get_children())

        if df.empty:
            return

        cols = list(df.columns)
        self.preview_tree["columns"] = cols
        for c in cols:
            self.preview_tree.heading(c, text=c)
            self.preview_tree.column(c, width=120, anchor="w")

        for _, row in df.iterrows():
            values = []
            for v in row.values:
                if isinstance(v, float):
                    values.append(f"{v:,.2f}")
                else:
                    values.append(str(v))
            self.preview_tree.insert("", "end", values=values)

    # ─────────────────────────────────────────────────────
    # Görselleştirme
    # ─────────────────────────────────────────────────────
    def show_visualization(self):
        try:
            df = self._get_dataframe()
        except Exception as e:
            messagebox.showerror("Veri Hatası", str(e))
            return

        self._populate_preview(df)

        if df.empty:
            messagebox.showinfo(
                "Veri Yok",
                "Bu rapor için henüz veri yok.\n"
                "Önce simülasyon çalıştırın veya canlı sipariş verin."
            )
            self._show_placeholder()
            return

        # Eski grafiği temizle, yeni figure'ı oluştur
        self.fig.clear()
        self.ax = self.fig.add_subplot(111)

        report = self.report_var.get()
        try:
            if report == "En Çok Satan Ürünler":
                self._plot_top_products(df)
            elif report == "Saatlik Ciro":
                self._plot_hourly_revenue(df)
            elif report == "Ürün × Saat Heatmap":
                self._plot_heatmap(df)
            elif report == "Fiyat Dağılımı (Violin)":
                self._plot_violin(df)
            elif report == "Kategori Cirosu":
                self._plot_category(df)
            elif report == "Genel Özet (KPI)":
                self._plot_kpi(df)
        except Exception as e:
            messagebox.showerror("Grafik Hatası", str(e))
            return

        self.fig.tight_layout()
        self.canvas.draw_idle()

    def _plot_top_products(self, df: pd.DataFrame):
        d = df.head(10)
        bars = self.ax.barh(d["Ürün"], d["Ciro"], color=config.COLOR_ACCENT)
        self.ax.set_xlabel("Ciro (₺)")
        self.ax.set_title("En Çok Satan Ürünler (Ciroya Göre)",
                          fontweight="bold")
        self.ax.invert_yaxis()
        for bar, val in zip(bars, d["Ciro"]):
            self.ax.text(val, bar.get_y() + bar.get_height()/2,
                         f"  ₺{val:,.0f}", va="center", fontsize=8)

    def _plot_hourly_revenue(self, df: pd.DataFrame):
        self.ax.bar(df["Saat"], df["Ciro"], color=config.COLOR_UP)
        self.ax.set_xlabel("Saat")
        self.ax.set_ylabel("Ciro (₺)")
        self.ax.set_title("Saatlik Ciro Dağılımı", fontweight="bold")
        self.ax.tick_params(axis="x", rotation=45)

    def _plot_heatmap(self, df: pd.DataFrame):
        # Pivot table → heatmap (seaborn)
        pivot = df.pivot_table(
            index="Ürün", columns="Saat", values="Ciro", aggfunc="sum", fill_value=0
        )
        sns.heatmap(
            pivot, ax=self.ax, cmap="YlOrRd",
            annot=True, fmt=".0f", cbar_kws={"label": "Ciro (₺)"},
            linewidths=0.3, linecolor="white"
        )
        self.ax.set_title("Ürün × Saat Ciro Heatmap (Seaborn)",
                          fontweight="bold")

    def _plot_violin(self, df: pd.DataFrame):
        # Çok ürün varsa en çok hareket edenleri al
        top_products = df["Ürün"].value_counts().head(8).index
        d = df[df["Ürün"].isin(top_products)]
        sns.violinplot(data=d, x="Ürün", y="Fiyat",
                       ax=self.ax, palette="Set2", inner="quartile")
        self.ax.set_title("Ürün Fiyat Dağılımları (Seaborn Violin)",
                          fontweight="bold")
        self.ax.tick_params(axis="x", rotation=30)

    def _plot_category(self, df: pd.DataFrame):
        colors = ["#3b82f6", "#16a34a", "#f59e0b", "#dc2626",
                  "#8b5cf6", "#06b6d4"]
        wedges, texts, autotexts = self.ax.pie(
            df["Ciro"], labels=df["Kategori"],
            autopct="%1.1f%%",
            colors=colors[:len(df)], startangle=90,
            textprops={"fontsize": 9},
        )
        for at in autotexts:
            at.set_color("white")
            at.set_fontweight("bold")
        self.ax.set_title("Kategori Bazında Ciro Dağılımı",
                          fontweight="bold")

    def _plot_kpi(self, df: pd.DataFrame):
        self.ax.axis("off")
        # Tablo şeklinde göster
        n = len(df)
        for i, row in df.iterrows():
            y = 0.9 - (i / max(n, 1)) * 0.8
            self.ax.text(0.05, y, row["Metrik"],
                         fontsize=12, fontweight="bold",
                         transform=self.ax.transAxes)
            self.ax.text(0.95, y, str(row["Değer"]),
                         fontsize=14, color=config.COLOR_ACCENT,
                         ha="right", transform=self.ax.transAxes)
        self.ax.set_title("Genel KPI Özeti", fontweight="bold")

    # ─────────────────────────────────────────────────────
    # Excel / CSV Export
    # ─────────────────────────────────────────────────────
    def export_excel(self):
        try:
            df = self._get_dataframe()
        except Exception as e:
            messagebox.showerror("Veri Hatası", str(e))
            return
        if df.empty:
            messagebox.showinfo("Veri Yok", "Aktarılacak veri yok.")
            return

        # Varsayılan dosya adı
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_name = self.report_var.get().replace(" ", "_")
        default_name = f"{report_name}_{ts}.xlsx"
        filepath = filedialog.asksaveasfilename(
            initialdir=config.EXPORT_DIR,
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not filepath:
            return

        try:
            with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="Rapor", index=False)
                # Başlığı kalın yap (basit format)
                ws = writer.sheets["Rapor"]
                from openpyxl.styles import Font, PatternFill
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill("solid", fgColor="3B82F6")
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
        except Exception as e:
            messagebox.showerror("Excel Hatası", str(e))
            return

        messagebox.showinfo("Tamam", f"Excel oluşturuldu:\n{filepath}")

    def export_csv(self):
        try:
            df = self._get_dataframe()
        except Exception as e:
            messagebox.showerror("Veri Hatası", str(e))
            return
        if df.empty:
            messagebox.showinfo("Veri Yok", "Aktarılacak veri yok.")
            return

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_name = self.report_var.get().replace(" ", "_")
        default_name = f"{report_name}_{ts}.csv"
        filepath = filedialog.asksaveasfilename(
            initialdir=config.EXPORT_DIR,
            initialfile=default_name,
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
        )
        if not filepath:
            return

        try:
            df.to_csv(filepath, index=False, encoding="utf-8-sig")  # Excel Türkçe için BOM
        except Exception as e:
            messagebox.showerror("CSV Hatası", str(e))
            return

        messagebox.showinfo("Tamam", f"CSV oluşturuldu:\n{filepath}")
